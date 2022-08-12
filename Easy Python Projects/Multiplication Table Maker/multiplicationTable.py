import sys
import openpyxl
from openpyxl.styles import Font


def multiplicationTable(n, filename='multiplicationTable.xlsx'):

    # create excel file
    wb = openpyxl.Workbook()

    # create worksheet
    sheet = wb.active
    sheet.title = '{}x{} multiplication table'.format(n, n)
    boldFont = Font(bold=True)
    
    # write row headers
    for i in range(1, n+1):
        sheet.cell(row=i+1, column=1).value = i
        sheet.cell(row=i+1, column=1).font = boldFont

    # write column headers
    for i in range(1, n+1):
        sheet.cell(row=1, column=i+1).value = i
        sheet.cell(row=1, column=i+1).font = boldFont

    # write multiplication table
    for row in range(1, n+1):
        for col in range(1, n+1):
            sheet.cell(row=row+1, column=col+1).value = row*col

    # save table
    wb.save(filename)


if __name__ == "__main__":

    if len(sys.argv) < 2:
        print('enter value for n')
    else:
        n = sys.argv[1]
        multiplicationTable(int(n))