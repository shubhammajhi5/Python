import os


import openpyxl


def textToSheet(directory, filename):
    
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title='result')
    sheet = wb.active

    colIndex = 1

    # write text files as columns in worksheet
    for file in os.listdir():
        if file.endswith('.txt'):
            rowIndex = 1
            with open(file) as f:
                for line in f:
                    sheet.cell(row=rowIndex, column=colIndex).value = line
                    rowIndex += 1
            colIndex += 1

    wb.save(filename)

if __name__ == "__main__":
    textToSheet('.', 'text-to-cols.xlsx')