import sys
import openpyxl


def blankRowInserter(index, num_blanks, filename):
    
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    rows = tuple(sheet.rows)

    
    for rowObj in rows[::-1]:
        for cellObj in rowObj:
            c = cellObj.column
            r = cellObj.row

            if r >= index and r < index+num_blanks:
                sheet.cell(row=r+num_blanks, column=c).value = cellObj.value
                sheet.cell(row=r, column=c).value = ''
            elif r >= index+num_blanks:
                sheet.cell(row=r+num_blanks, column=c).value = cellObj.value
    
    wb.save('result_'+filename)


if __name__ == "__main__":
    
    num_args = len(sys.argv)

    if num_args < 4:
        print("usage: python blankRowInserter.py <index> <no_of_blank_rows> filename.xlsx")
    else:
        index = int(sys.argv[1])
        num_blanks = int(sys.argv[2])
        filename = sys.argv[3]

        blankRowInserter(index, num_blanks, filename)